# openpyxl - para trabalhar com arquivos Excel xlsx, xlsm, xltx e xltm
# Com essa biblioteca será possível ler e escrever dados em células
# especificas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envonveldo planilhas do 
# Excel, com a criação de relatórios e análise de dadose/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Doc: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
worksheet = Worksheet = workbook.active

# Criando os cabeçalhos da planilha
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # nome       idade nota
    ['Alexandre', 20, 5.5],
    ['Maria',     13, 9.5],
    ['Valéria',   15, 7.5],
    ['Paulo',     16, 6.5],
]

#for i in range(2, 10):
#    for j in range(1, 4):
#        print('Linha', i, 'Coluna', j)

''' Melhor forma de criar a planilha

for i, student_row in enumerate(students, start=2):
    for j, student_column in enumerate(student_row, start=1):
        worksheet.cell(i, j, student_column)'''

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)