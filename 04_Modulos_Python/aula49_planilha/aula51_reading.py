# openpyxl - ler a alterar dadoss de uma planilha
# Com essa biblioteca será possível ler e escrever dados em células
# especificas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envonveldo planilhas do 
# Excel, com a criação de relatórios e análise de dadose/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Doc: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregando um arquivo do excel
workbook: Workbook = load_workbook(WORKBOOK_PATH)

sheet_name = 'Minha planilha' # nome da planilha
worksheet = Worksheet = workbook[sheet_name] # selecionando a planilha
max_lengths = [0, 0, 0] # definindo o comprimento máximo para cada coluna

row: tuple[Cell]
for row in worksheet.iter_rows():
    for i, cell in enumerate(row):
        max_lengths[i] = max(max_lengths[i], len(str(cell.value)))

        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2, 19)

# Imprimir a tabela com espaçamento correto
for row in worksheet.iter_rows(values_only=True):
    for i, cell in enumerate(row):
        print(str(cell).ljust(max_lengths[i] + 2), end='')
    print()

'''print(f'{worksheet['B1'].value} anterior da {worksheet["A3"].value}: {worksheet['B3'].value}')
worksheet['B3'].value = 18
print(f'{worksheet['B1'].value} atual da {worksheet["A3"].value}: {worksheet['B3'].value}')'''


workbook.save(WORKBOOK_PATH)